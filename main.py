"""
    Imported libraries used for the code
"""
import openpyxl
import dllist
import networkx as nx
from kivy.app import App
from kivy.uix.screenmanager import ScreenManager
from kivy.config import Config
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.popup import Popup
from datetime import datetime, timedelta

"""
    Main class holding the logical part
"""


class TubeScreen(ScreenManager):
    """
        Initial Variables of the class
    """
    wb = openpyxl.load_workbook(filename='London Underground data.xlsx')
    sh = wb.active
    dll = dllist.DoublyLinkedList()
    G = nx.Graph()

    def spinner_filler(self):
        """
            Function tha fill the spinner objects with the data
        """
        filler = []
        for i in range(1, self.sh.max_row + 1):
            if self.sh.cell(row=i, column=2).value is not None or 0:
                filler.append(self.sh.cell(row=i, column=2).value)
        final_filler = set(filler)
        self.get_data_for_dll()
        return sorted(final_filler)

    def calculate_route(self):
        """
            Function that will calculate the route
        """
        connect = self.dll.return_data_for_alg()
        for i in range(len(connect)):
            self.G.add_weighted_edges_from(connect)
        route = nx.dijkstra_path(self.G, self.ids.spinner1.text, self.ids.spinner2.text)
        return route

    def time(self, full_info):
        """
            Function that takes care of departure and arrival time
        """
        journey_time = 0
        for x in range(len(full_info)):
            journey_time = journey_time + full_info[x][3]

        time = datetime.now()
        arrival = datetime.now() + timedelta(minutes=journey_time)
        self.ids.start.text = 'Starting time:                       ' + time.strftime('%H:%M')
        self.ids.end.text = 'Estimated time of arrival:   ' + arrival.strftime('%H:%M')

    def repack(self, route):
        """
            Function that repacks route in a list of lists
        """
        route_col1 = []
        route_col2 = []
        path = []
        for z in range(1, len(route)):
            route_col1.append(route[z])

        for z in range(0, len(route) - 1):
            route_col2.append(route[z])

        for z in range(0, len(route_col2)):
            path.append([route_col2[z], route_col1[z]])
        return path

    def get_data_for_dll(self):
        """
             Feed the doubly linked list
        """
        for i in range(1, self.sh.max_row + 1):
            if self.sh.cell(row=i, column=2).value and self.sh.cell(row=i, column=3).value \
                    and self.sh.cell(row=i, column=4).value is not None:
                self.dll.append([self.sh.cell(row=i, column=1).value,
                                 self.sh.cell(row=i, column=2).value,
                                 self.sh.cell(row=i, column=3).value,
                                 self.sh.cell(row=i, column=4).value])

    def get_full_route_detail(self):
        """
             Get all details related to the route form the doubly linked list
        """
        full_info = []
        path = self.repack(self.calculate_route())
        connect = self.dll.return_all_data()
        for q in range(len(path)):
            for j in range(len(connect)):
                if path[q][0] == connect[j][2] and path[q][1] == connect[j][1] \
                        or path[q][0] == connect[j][1] and path[q][1] == connect[j][2]:
                    full_info.append(connect[j])
        return full_info

    def reverse(self, list):
        """
           Reverse lists
        """
        lst = list
        reveser = [lst.pop(1), lst.pop(1)]
        reveser.reverse()
        lst.insert(1, reveser[1])
        lst.insert(1, reveser[0])
        return lst

    def reverse_if_needed(self):
        """
        Reverse when needed
        """
        final = []
        full_info = self.eliminate_repetition()
        for i in range(1, len(full_info)):
            if full_info[i - 1][2] == full_info[i][1]:
                final.append(full_info[i])
            elif full_info[i - 1][2] != full_info[i][1]:
                final.append(self.reverse(full_info[i]))

        return final

    def eliminate_repetition(self):
        """
            Eliminate unnecessary repetitions
        """
        full_info = self.get_full_route_detail()
        new_full_info = []
        for x in range(len(full_info)):
            if full_info[x][1] != full_info[x - 1][1] or full_info[x][2] != full_info[x - 1][2]:
                new_full_info.append(full_info[x])
        return new_full_info

    def get_changes(self):
        """
            Function to detect changes
        """
        full_info = self.reverse_if_needed()
        change = []
        display_change = []
        if len(full_info) == 1:
            change.append(['At ', full_info[0][1], ' change to ' + full_info[0][0] + ' Line'])
        for d in range(len(full_info)):
            if full_info[d - 1][0] != full_info[d][0] and full_info[d - 1][2] == full_info[d][1]:
                change.append(['At ', full_info[d][1], ' change to ' + full_info[d][0] + ' Line'])
        for j in range(len(change)):
            if not change[j] in display_change:
                display_change.append(change[j])
        return display_change

    def generate_widgets(self):
        """
            Dynamically generate label widgets for layouts to display the route and changes
        """
        route = self.calculate_route()
        change = self.get_changes()
        for f in range(len(change)):
            self.ids.change.add_widget(Label(text=str(change[f]), color=[0, 0, 0], text_size=[350, 0],
                                             halign='left', valign='top', ))
            self.ids.change.height += 40
        for f in range(len(route)):
            self.ids.s_layout.add_widget(Label(text=route[f], color=[0, 0, 0], text_size=[350, 0],
                                               halign='left', valign='top', ))
            self.ids.s_layout.height += 40

    def set_label_text(self):
        """
            Set the tex for one specific label
        """
        self.ids.lbchanges.text = ' Changes: '

    def clean_layouts(self):
        """
            clean the layouts for a new search
        """
        self.ids.s_layout.clear_widgets()
        self.ids.change.clear_widgets()
        self.ids.change.height = 0
        self.ids.s_layout.height = 0

    def popup_error(self):
        """
            Displays a popup widget with an error message
        """
        show = FloatLayout()
        txt_pop = Label(text='Please, Select your Departure and Destination',
                        pos_hint={'x': .01, 'y': .20})
        pop_error = Popup(title='Error',
                          size_hint=(None, None),
                          size=(400, 400),
                          content=show
                          )
        bnt_pop = Button(text='exit', size_hint=(None, None), size=(300, 50),
                         on_release=pop_error.dismiss,
                         pos_hint={'x': .1, 'y': .1})
        show.add_widget(txt_pop)
        show.add_widget(bnt_pop)

        return pop_error.open()

    def starter(self):
        """
            Call most of the functions to make the search
        """
        if self.ids.spinner1.text == 'Please, Where From?' or self.ids.spinner2.text == 'Please, Where To?':
            self.ids.btn1.disable = True

            return self.popup_error()
        else:
            self.ids.btn1.disable = False
            self.clean_layouts()
            self.time(self.reverse_if_needed())
            self.generate_widgets()
            self.set_label_text()


class TubeApp(App):
    """
    Class tha take care of core Kivy application
    """
    title = 'TubeApp'
    Config.set('graphics', 'window_state', 'maximized')

    def build(self):
        return TubeScreen()


TubeApp().run()
